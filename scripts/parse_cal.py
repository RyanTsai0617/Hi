"""Parse Bloomberg economic calendar Excel and dump rows for filtering."""
from openpyxl import load_workbook
import json
from datetime import datetime, date
import os

PATH = r"C:\Users\tsait\OneDrive\桌面\04_週報日報\每周報告\20260420\經濟數據\27152250_20260418_081357_eco.xlsx"

wb = load_workbook(PATH, data_only=True)
for sn in wb.sheetnames:
    ws = wb[sn]
    print(f"Sheet: {sn} | rows={ws.max_row} cols={ws.max_column}")
    headers = [c.value for c in ws[1]]
    print("Headers:", headers)
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=4, values_only=True), start=2):
        print(f"Row {i}:", row)
    print()

ws = wb[wb.sheetnames[0]]
headers = [c.value for c in ws[1]]
rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    rec = {}
    for h, v in zip(headers, row):
        if isinstance(v, (datetime, date)):
            v = v.isoformat()
        rec[str(h)] = v
    rows.append(rec)

out = r"C:\Users\tsait\OneDrive\桌面\Coding\.tmp\eco_rows.json"
os.makedirs(os.path.dirname(out), exist_ok=True)
with open(out, "w", encoding="utf-8") as f:
    json.dump(rows, f, ensure_ascii=False, indent=2, default=str)
print(f"Dumped {len(rows)} rows to {out}")
