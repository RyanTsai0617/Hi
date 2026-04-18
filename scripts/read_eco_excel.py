"""Inspect Bloomberg economic calendar Excel and dump rows for parsing."""
from openpyxl import load_workbook
import json
from datetime import datetime, date

PATH = r"C:\Users\tsait\OneDrive\桌面\04_週報日報\每周報告\20260420\經濟數據\27152250_20260418_081357_eco.xlsx"

wb = load_workbook(PATH, data_only=True)
for sn in wb.sheetnames:
    ws = wb[sn]
    print(f"=== Sheet: {sn} | rows={ws.max_row} cols={ws.max_column} ===")
    # Print header
    headers = [c.value for c in ws[1]]
    print("Headers:", headers)
    # Print first 5 rows
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=6, values_only=True), start=2):
        print(f"Row {i}:", row)
    print()

# Also dump all rows from first sheet to stdout as JSON so we can parse
ws = wb[wb.sheetnames[0]]
headers = [c.value for c in ws[1]]
rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    rec = {}
    for h, v in zip(headers, row):
        if isinstance(v, (datetime, date)):
            v = v.isoformat()
        rec[str(h)] = v
    rows.append(rec)

out = r"C:\Users\tsait\OneDrive\桌面\Coding\.tmp\eco_rows.json"
import os
os.makedirs(os.path.dirname(out), exist_ok=True)
with open(out, "w", encoding="utf-8") as f:
    json.dump(rows, f, ensure_ascii=False, indent=2, default=str)
print(f"Dumped {len(rows)} rows to {out}")
